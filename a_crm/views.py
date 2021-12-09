from django.contrib.auth.decorators import login_required
from django.forms.widgets import NullBooleanSelect
from django.http.response import HttpResponse, HttpResponseForbidden, HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls.base import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, DeleteView, TemplateView, UpdateView 
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db.models import *

from .models import *
from .forms import *
from .utils import *

import openpyxl
from datetime import datetime


@login_required (login_url='/lgn')
def invoice(request, itm_id):
    
    blank = settings.MEDIA_ROOT + '/invoices/' + 'blank.xlsx' 
    savet = settings.MEDIA_ROOT + '/invoices/invoice_' + str(itm_id) + '.xlsx'
    link  = 'media/invoices/invoice_' + str(itm_id) + '.xlsx'
    data = Items.objects.get(id=itm_id)

    rd = openpyxl.load_workbook(blank)
    sheet = rd.active

    sheet.cell(16, 11).value = itm_id
    sheet.cell(16, 15).value = datetime.now().date()
    sheet.cell(22, 3).value = 'Портрет ум. ' + str(data.surn) + ' ' + str(data.name) + ' ' + str(data.midn)
    sheet.cell(22, 7).value = str(data.code)
    sheet.cell(45, 7).value = datetime.now().date()

    sheet.cell(22, 24).value = data.code.cost

    sheet.cell(22, 26).value = data.code.cost * 1
    sheet.cell(23, 26).value = data.code.cost * 1
    sheet.cell(24, 26).value = data.code.cost * 1

    sheet.cell(22, 37).value = data.code.cost * 1
    sheet.cell(23, 37).value = data.code.cost * 1
    sheet.cell(24, 37).value = data.code.cost * 1

    sheet.cell(36, 2).value = data.code.word

    rd.save(savet)

    context = {
        'title': 'Накладная по заказу № ' + str(itm_id),
        'link': link,
    }

    return render(request, 'a_crm/invoice.html', context)


@login_required(login_url='/lgn')
def che(request, itm_id):

    data = Items.objects.get(id = itm_id)
    datalist = data.__dict__

    if request.method == 'POST':
        form = ItmForm(request.POST, request.FILES)
        if form.is_valid():
            # create, but not save new instance
            saving = form.save(commit=False) 

            # setting file values 
            if not request.FILES.get('phot_natu'):
                saving.phot_natu = data.phot_natu
            if not request.FILES.get('phot_reto'):
                saving.phot_reto = data.phot_reto

            # setting manager values 
            if request.POST.get('mana_chec'):
                mana = User.objects.get(id=request.user.pk)
            else:
                mana = None

            saving.mana = mana
            saving.id = itm_id
            saving.auth = data.auth
            saving.time_crea = data.time_crea
            saving.save()
            
            return redirect('che', itm_id=itm_id)
    else:
        datalist = data.__dict__
        datalist['phot_natu'] = data.phot_natu
        datalist['phot_reto'] = data.phot_reto
        if data.mana:
            datalist['mana_chec'] = True
        else:
            datalist['mana_chec'] = False

        form = ItmForm(initial=datalist)

    context = {
        'title': 'Редактор заказа',
        'form': form,
        'itm_id': itm_id,
        'data': data,
        'datalist': datalist,
    }

    return render (request, 'a_crm/che.html', context)


class Add(LoginRequiredMixin, CreateView):
    form_class = ItmForm
    template_name = 'a_crm/add.html'
    success_url = reverse_lazy('lst')
    login_url = reverse_lazy('lgn')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Разместить заказ на изготовление таблички для умершего"
        return context

    def form_valid(self, form):
        form.instance.auth = self.request.user
        return super().form_valid(form)


class Itm(LoginRequiredMixin, DetailView):
    model =Items
    template_name = 'a_crm/itm.html'
    pk_url_kwarg = 'itm_id'
    context_object_name = 'info'
    login_url = reverse_lazy('lgn')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = str(context['info'].name) + ' ' + str(context['info'].surn) + ' ' + str(context['info'].midn)
        context['auth'] = User.objects.get(username=context['info'].auth).last_name
        if context['info'].mana:
            context['mana'] = User.objects.get(username=context['info'].mana).last_name
        else:
            context['mana'] = 'не назаначен'
        return context


class Rem(LoginRequiredMixin, IsmaMixin, DeleteView):
    model = Items
    template_name = 'a_crm/rem.html'
    pk_url_kwarg = 'itm_id'
    success_url = reverse_lazy('lst')
    login_url = reverse_lazy('lgn')

    def delete(self, request, *args, **kwargs):
        if self.is_ma():
            self.object = self.get_object()
            success_url = self.get_success_url()
            self.object.delete()
            return HttpResponseRedirect(success_url)
        else:
            return HttpResponseForbidden('<h1>Вам нельзя этого делать</h1>')


class Lst(LoginRequiredMixin, IsmaMixin, ListView):
    model = Items
    template_name = 'a_crm/lst.html'
    context_object_name = 'data'
    login_url = reverse_lazy('lgn')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Все заказы'
        context['len'] = len(self.get_queryset())
        context['meth'] = self.request.method
        context['searform'] = self.get_form()
        return context

    def get_form(self):
        if self.request.method == "POST":
            searform = SearForm(self.request.POST)
        else:
            searform = SearForm()
        return searform

    def post(self, request, *args, **kwargs):
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        self.get_queryset = Items.objects.all()
        return self.render_to_response(context)
    
    def get_queryset(self):
        if self.is_ma():
            q = Items.objects.all()
        else:
            q = Items.objects.filter(auth=self.request.user)

        if self.request.method == "POST":
            agent = self.request.POST.get('agent')
            if agent and len(agent) > 2:
                u = User.objects.filter(Q(first_name__contains=agent) | Q(last_name__contains=agent))
                if len(u) > 0:
                    q_list = []
                    for i in u:
                        q_list.append(i.pk)
                    q = q.filter(auth__in=q_list)
                else:
                    q = []
            else:
                q = []


        fini = self.request.GET.get('fini')
        if fini:
            return q.filter(fini=fini)
        else:
            return q


def test(request, bla):
    data = request
    context = {
        'data': data.GET,
        'bla': bla,
    }
    print(bla)
    return render(request, 'a_crm/test.html', context)












# def che(request, itm_id):

#     data = Items.objects.get(id = itm_id)

#     form = ItmForm(initial=data.__dict__)

#     context = {
#         'title': 'title',
#         'itm_id': itm_id,
#         'form': form,
#         'data': data
#     }

#     return render (request, 'a_crm/che.html', context)






# def che(request, itm_id):

#     grou = []
#     group = Group.objects.filter(user = request.user)
#     for g in group:
#         grou.append(g.name)

#     data = Items.objects.get(id = itm_id)

#     # auth = User.objects.get(id=data.user.id)

#     if request.method == 'POST':

#         form = ItmFormChenMan(request.POST, request.FILES)
#         if form.is_valid():
#             saving = form.save(commit=False)

#             phot_natu = request.FILES.get('phot_natu')
#             phot_reto = request.FILES.get('phot_reto')


#             if not phot_natu:
#                 saving.phot_natu = data.phot_natu
#             if not phot_reto:
#                 saving.phot_reto = data.phot_reto
            
#             saving.id = itm_id
#             saving.auth = data.auth
#             saving.mana = request.user
#             saving.time_crea = data.time_crea
#             saving.save()
#             return redirect('itm', itm_id=itm_id)
#     else:
#         if 'Менеджеры' in grou:
#             datalist = {
#                 'phot_with': data.phot_with, 
#                 'surn': data.surn, 
#                 'name': data.name, 
#                 'midn': data.midn, 
#                 'date_birt': data.date_birt, 
#                 'date_deat': data.date_deat, 
#                 'date_fune': data.date_fune, 
#                 'code': data.code, 
#                 'numb': data.numb, 
#                 'colo': data.colo, 
#                 'fast': data.fast, 
#                 'posi': data.posi, 
#                 'addr': data.addr, 
#                 'phot_natu': data.phot_natu, 
#                 'phot_reto': data.phot_reto, 
#                 'desc': data.desc,
#                 'agre': data.agre, 
#                 'fini': data.fini,
#                 'auth': data.auth,
#             } 
#             form = ItmFormChenMan(initial=datalist)          
#             # form = ItmFormChenMan(data)          
#         else:
#             datalist = {
#                 'phot_with': data.phot_with, 
#                 'surn': data.surn, 
#                 'name': data.name, 
#                 'midn': data.midn, 
#                 'date_birt': data.date_birt, 
#                 'date_deat': data.date_deat, 
#                 'date_fune': data.date_fune, 
#                 'code': data.code, 
#                 'numb': data.numb, 
#                 'colo': data.colo, 
#                 'fast': data.fast, 
#                 'posi': data.posi, 
#                 'addr': data.addr, 
#                 'phot_natu': data.phot_natu, 
#                 'desc': data.desc,
#                 'auth': data.auth,
#             }
#             form = ItmFormChenAge(initial=datalist)
#             # form = ItmFormChenAge(data)

#     context = {
#         'title': 'Редактор заказа',
#         'form': form,
#         'grou': grou,
#         'itm_id': itm_id,
#     }

#     return render (request, 'a_crm/che.html', context)







# @login_required(login_url='/lgn')
# def lst(request):

#     user = request.user

#     grou = []
#     group = Group.objects.filter(user = user)
#     for g in group:
#         grou.append(g.name)

#     fini = request.GET.get('fini')
#     data = Items.objects.all()

#     if fini == None:
#         if 'Менеджеры' in grou:
#             data = data
#         else:
#             data = data.filter(auth=user)
#     else:
#         if 'Менеджеры' in grou:
#             data = data.filter(fini=fini)
#         else:
#             data = data.filter(auth=user, fini=fini)

#     context = {
#         'title': 'Заказы',
#         'data': data,
#         'grou': grou,
#         'user': user,
#         'fini': fini,
#     }

#     return render (request, 'a_crm/lst.html', context)




# class Che(LoginRequiredMixin, UpdateView):

#     model = Items
#     form_class = ItmForm
#     pk_url_kwarg = 'itm_id'
#     template_name = 'a_crm/che.html'
#     success_url = reverse_lazy('home')
#     login_url = reverse_lazy('lgn')

#     def get_context_data(self, *, object_list=None, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['title'] = "Редактор заказа"
#         context['id'] = self.kwargs.get(self.pk_url_kwarg)
#         return context
